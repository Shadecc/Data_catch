# coding=gbk
import io
import sys
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import time

# aim_row�У�aim_col��
aim_col = 2
aim_row = 1
file_name = 'data.xlsx'
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='gb18030') #�ı��׼�����Ĭ�ϱ���
browser = webdriver.Firefox()
wait = WebDriverWait(browser, 10)


def index_page(i):
    """
    ����ÿһҳ����
    :param i: ҳ��ĵ� i ҳ
    """
    if i == 1:
        url = "http://star.iecity.com/all/0"
        browser.get(url)

    # �ȴ� Main �ڵ���س���
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#Main')))

    # ���� get_info() ������ҳ����н���
    get_info()
    # Ѱ����һҳ����Ľڵ�
    next_p = browser.find_element(By.XPATH, ('//*[@id="Main"]/div[4]/div/span[12]'))
    # �ҵ���ͣ�� 30 ��
    time.sleep(4)
    # �����ť
    next_p.click()


def main():
    """
    ����ҳ��
    :return:
    """
    for i in range(1, 354):
        index_page(i)


def get_info():
    """
    ��ȡÿҳ����
    #
    :return:
    """
    # �ҵ��½ڵ�����
    content = browser.find_element_by_xpath('/html/body/div[3]/div[2]/div[4]/ul')
    li_list = content.find_elements_by_tag_name('li')
    global aim_row
    for li in li_list:
        li_content = li.text
        aim_row += 1
        # ����ʵ��aim_row�У�aim_col���޸����ݼ���
        # �����Ѵ��ڵ�wookbook����
        wb = load_workbook(file_name)
        wb1 = wb.active  # ����sheet
        # ��sheet��д������
        wb1.cell(aim_row, 1, aim_row-1)
        wb1.cell(aim_row, aim_col, li_content)
        # ����
        wb.save(file_name)


if __name__ == '__main__':
    main()